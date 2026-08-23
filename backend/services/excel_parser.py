"""Excel file parser service."""
import openpyxl
from io import BytesIO
from typing import Any


def parse_excel(file_bytes: bytes) -> dict[str, Any]:
    """
    Parse an Excel file and extract column headers + preview rows.
    
    Returns:
        {
            "columns": ["Name", "Date", "Amount", ...],
            "preview": [
                {"Name": "John Doe", "Date": "2024-01-15", "Amount": 5000},
                ...
            ],
            "total_rows": 100
        }
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Extract headers from first row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        val = cell.value
        if val is not None:
            headers.append(str(val).strip())
        else:
            break  # Stop at first empty column

    if not headers:
        raise ValueError("Excel file has no column headers in the first row.")

    # Extract preview rows (first 5 data rows)
    preview = []
    total_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        if len(preview) < 5:
            row_data = {}
            for i, header in enumerate(headers):
                val = row[i] if i < len(row) else None
                row_data[header] = str(val) if val is not None else ""
            preview.append(row_data)

    wb.close()

    return {
        "columns": headers,
        "preview": preview,
        "total_rows": total_rows,
    }


def get_row_data(file_bytes: bytes, row_index: int = 0) -> dict[str, str]:
    """
    Get data for a specific row (0-indexed from data rows, not header).
    Returns a dict mapping column header -> cell value as string.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        val = cell.value
        if val is not None:
            headers.append(str(val).strip())
        else:
            break

    target_row = row_index + 2  # +1 for header, +1 for 1-indexed
    row_data = {}
    for row in ws.iter_rows(min_row=target_row, max_row=target_row, values_only=True):
        for i, header in enumerate(headers):
            val = row[i] if i < len(row) else None
            row_data[header] = str(val) if val is not None else ""
        break

    wb.close()
    return row_data
